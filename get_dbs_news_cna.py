from urllib.request import urlopen
from urllib.error import HTTPError
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import pyautogui as pag
import os
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime as dt
from datetime import timedelta as time_delta
#import xlsxwriter
#import pygetwindow
#from MasterScript import daily, toPandas, howManyData, tailNum
from selenium.webdriver.chrome.options import Options
import pywhatkit as pwk
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import emails
import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import re

options = Options()
options.add_experimental_option("detach", True)
#options.add_experimental_option("excludeSwitches", ['enable-automation'])


url = "https://www.channelnewsasia.com/search?q=dbs"

try:
    page = requests.get(url)
except HTTPError as ex:
    print("URL is not valid, please check URL address again.")
    print (ex)
else:
    bs_obj = BeautifulSoup(page.content, 'html.parser')
    print("ok")
    #print(bs_obj)

# #Open the website with chrome webdriver
driver_path = r'C:\Users\deadw\AppData\Local\Programs\Python\Python310\Scripts\chromedriver.exe'
driver = webdriver.Chrome(chrome_options = options, executable_path = driver_path)
driver.get(url)

# pagedown to download all data
a = ActionChains(driver)
for i in range(0,5,1):
    #perform the ctrl+end pressing action
    a.key_down(Keys.CONTROL).key_down(Keys.END).key_up(Keys.CONTROL).key_up(Keys.END).perform()
    time.sleep(1)



pageSource = driver.page_source
bsObj_1 = BeautifulSoup(pageSource, 'html.parser')

#section = bsObj_1.find("section", {"id": "block-mc-cna-theme-mainpagecontent"}).text

section = bsObj_1.find("div", {"class": "top-stories-primary-section__items top-stories-primary-section__items--col-one"}).text
section = re.sub(r'\n\s*\n', r'\n\n', section.strip(), flags=re.M)
section_2 = bsObj_1.find("div", {"class": "top-stories-primary-section__items top-stories-primary-section__items--col-one"})
print(section_2)
#Get all href and put into a list
header_list = []
time_list = []
href_list = []
for link in section_2.find_all('a'):
    news_link = "https://www.channelnewsasia.com" + (link.get('href'))
    href_list.append(news_link)
    #drop duplicates
    href_list = list(dict.fromkeys(href_list))
print(href_list)

for header_link in section_2.find_all('img'):
    title = header_link.get('alt')
    header_list.append(title)
print(header_list)

for time_link in section_2.find_all('span',{"class":"list-object__timestamp list-object__timestamp-- timestamp timeago"}):
    ttime = time_link.text
    time_list.append(ttime)
print(time_list)

for k in range(0, len(header_list)):
    print(header_list[k])
    print(time_list[k])
    print(href_list[k])
    print("")
    
#Create the date
today_date = dt.today().date()
#Create the time
today_time = dt.today().time()

#Group data to list elements in a master list
all_data_list = []
for j in range(0, len(header_list)):
    all_data_list.append(today_date)
    all_data_list.append("cna")
    all_data_list.append(header_list[j])
    all_data_list.append(header_list[j])
    all_data_list.append(href_list[j])
    all_data_list.append("business")
    all_data_list.append(time_list[j])
    all_data_list.append(today_time)

#Split all_data_list into 2d list
final_list = [all_data_list[i:i+8] for i in range(0, len(all_data_list), 8)]
print(final_list)


#load list to pandas df
df = pd.DataFrame (final_list, columns = ["date", "source", "header", "summary", "link", "category", "timeline", "time"])
print(df)


#pandas df to Excel sheet
#pandas df to Excel sheet
excel_file = r'C:\Users\deadw\Documents\Algo\May2021\News\business_news_cna.xlsx'
#get last empty row
wb = load_workbook(excel_file)
sheet = wb.active
last_empty_row = len(list(sheet.rows)) + 1
print('last empty row in Excel : ' + str(last_empty_row))
wb.close()

#load df to Excel 
wb = load_workbook(excel_file)
sheet = wb.active
r = last_empty_row
c = 0
df_row_count = df.shape[0]  # Gives number of rows
print('number of rows in df = ' + str(df_row_count))
for i in range(0, df_row_count):
    sheet.cell(row = r, column = 1).value = df['date'].iloc[i]
    sheet.cell(row = r, column = 2).value = df['source'].iloc[i]
    sheet.cell(row = r, column = 3).value = df['header'].iloc[i]
    sheet.cell(row = r, column = 4).value = df['summary'].iloc[i]
    sheet.cell(row = r, column = 5).value = df['link'].iloc[i]
    sheet.cell(row = r, column = 6).value = df['category'].iloc[i]
    sheet.cell(row = r, column = 7).value = df['timeline'].iloc[i]
    sheet.cell(row = r, column = 8).value = df['time'].iloc[i]
    r = r + 1

wb.save(excel_file)
driver.close()



