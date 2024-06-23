from urllib.request import urlopen
from urllib.error import HTTPError
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.ui import Select
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import pyautogui as pag
import os
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime as dt
from datetime import timedelta as time_delta
#import xlsxwriter
#import pygetwindow
#from MasterScript import daily, toPandas, howManyData, tailNum
from selenium.webdriver.chrome.options import Options
import pywhatkit as pwk
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import emails
import email, smtplib, ssl

from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

options = Options()
options.add_experimental_option("detach", True)
#options.add_experimental_option("excludeSwitches", ['enable-automation'])


url = "https://www.marketwatch.com/column/the-fed"
excel_file = r'C:\Users\deadw\Documents\Algo\May2021\News\fed_news.xlsx'

try:
    page = requests.get(url)
except HTTPError as ex:
    print("URL is not valid, please check URL address again.")
    print (ex)
else:
    bs_obj = BeautifulSoup(page.content, 'html.parser')
    print("ok")
    #print(bs_obj)

# #Open the website with chrome webdriver
driver_path = r'C:\Users\deadw\AppData\Local\Programs\Python\Python310\Scripts\chromedriver.exe'
driver = webdriver.Chrome(chrome_options = options, executable_path = driver_path)
driver.get(url)

# pagedown to download all data
a = ActionChains(driver)
for i in range(0,10,1):
    #perform the ctrl+end pressing action
    a.key_down(Keys.CONTROL).key_down(Keys.END).key_up(Keys.CONTROL).key_up(Keys.END).perform()
    time.sleep(1)



pageSource = driver.page_source
bsObj_1 = BeautifulSoup(pageSource, 'html.parser')

section = bsObj_1.find("div", {"id": "Main"})

sub_sections = section.find_all("li", {"class": "js-stream-content Pos(r)"})
news_list = []
for ss in sub_sections:
    link = ss.find('a')
    if ss.text[0] != "A" and ss.text[1] != "d":
        #all_span = sub_sections.get_all('span')
        #for sp in all_span:
             #print(sp.text)
        #soup.find("div", class_="sr-2").find_all("span")[1].text
        #sp_1 = ss.select('span')[0]
        #sp_2 = ss.select('span')[1]
        #(soup.select("div.sr-2 > span")[1].text)  
        #sp_2 = ss.select('div.Mx')('span')[0]
        #select('div.line')[2].select('span')[1] 
        category = ss.select('div')[1]
        for cat in category:
            for cat1 in cat:
        # sp_1 = category.find('span')[0]
        # header = category.find('h3')
        # para = category.find('p')
                print(cat1.text)
                news_list.append(cat1.text)

        # print(category.text)    
        # print(sp_1.text)

        # #print(sp_2.text)
        # print(header.text)
        # print(para.text)
        cat_link = 'https://sg.finance.yahoo.com'+link.attrs['href']
        print(cat_link)
        news_list.append(cat_link)
        
        print(" ")
        print(" ")


#removes empty string in news_list
news_list = [x for x in news_list if x]
#drop first 2 elements
n = 3
del news_list[:n]
#print news_list
#print(news_list)
#print(*news_list, sep = ", ") 


#Put each match details into a list
final_list = [news_list[i:i+5] for i in range(0, len(news_list), 5)]
print(final_list)


#load list to pandas df
df = pd.DataFrame (final_list, columns = ["source", "header", "summary", "link", "category"])
print(df)

#Split the source and date 
df[['source','timeline']] = df['source'].str.split('•',expand=True)

#Create the date column
#Create the date column
today_date = dt.today()
yesterday_date = today_date - time_delta(1)
two_days_ago = today_date - time_delta(2)
three_days_ago = today_date - time_delta(3)
today_date = today_date.strftime("%Y/%m/%d")
yesterday_date = yesterday_date.strftime("%Y/%m/%d")
two_days_ago = two_days_ago.strftime("%Y/%m/%d")
three_days_ago = three_days_ago.strftime("%Y/%m/%d")

df['date'] = pd.np.where(df.timeline.str.contains("hours ago"), today_date,
                            pd.np.where(df.timeline.str.contains("today"), today_date,
                            pd.np.where(df.timeline.str.contains("hour ago"), today_date,
                            pd.np.where(df.timeline.str.contains("minutes"), today_date,
                            pd.np.where(df.timeline.str.contains("yesterday"), yesterday_date,
                            pd.np.where(df.timeline.str.contains("2 days ago"), two_days_ago,
                            pd.np.where(df.timeline.str.contains("3 days ago"), three_days_ago, today_date)))))))

#move date column to first column
col = df.pop("date")
df.insert(0, col.name, col)

#drop row if any column is blank
df.dropna(
    axis=0,
    how='any',
    thresh=None,
    subset=None,
    inplace=True
)

#pandas df to Excel sheet
#pandas df to Excel sheet
excel_file = r'C:\Users\deadw\Documents\Algo\May2021\News\daily_news.xlsx'
#get last empty row
wb = load_workbook(excel_file)
sheet = wb.active
last_empty_row = len(list(sheet.rows)) + 1
print('last empty row in Excel : ' + str(last_empty_row))
wb.close()

#load df to Excel 
wb = load_workbook(excel_file)
sheet = wb.active
r = last_empty_row
c = 0
df_row_count = df.shape[0]  # Gives number of rows
print('number of rows in df = ' + str(df_row_count))
for i in range(0, df_row_count):
    sheet.cell(row = r, column = 1).value = df['date'].iloc[i]
    sheet.cell(row = r, column = 2).value = df['source'].iloc[i]
    sheet.cell(row = r, column = 3).value = df['header'].iloc[i]
    sheet.cell(row = r, column = 4).value = df['summary'].iloc[i]
    sheet.cell(row = r, column = 5).value = df['link'].iloc[i]
    sheet.cell(row = r, column = 6).value = df['category'].iloc[i]
    sheet.cell(row = r, column = 7).value = df['timeline'].iloc[i]
    r = r + 1

wb.save(excel_file)

#dataframe to html table
html_table = df.to_html()
#pwk.sendwhatmsg('+6596905822', 'testing 123', 7, 21)

#send email with html table
# Here goes the configuration of your email provider.
# Look online to find it.
# For instance for Yahoo!Mail:

subject = "News from news"
body = "This is an email with news attachment sent from Python"
sender_email = "chinchaiseng@gmail.com"
receiver_email = "chinchaiseng@gmail.com"
password = input("Type your password and press enter:")

# Create a multipart message and set headers
message = MIMEMultipart()
message["From"] = sender_email
message["To"] = receiver_email
message["Subject"] = subject
message["Bcc"] = receiver_email  # Recommended for mass emails

# Add body to email
message.attach(MIMEText(body, "plain"))

filename = "daily_news.xlsx"  # In same directory as script

# Open PDF file in binary mode
with open(filename, "rb") as attachment:
    # Add file as application/octet-stream
    # Email client can usually download this automatically as attachment
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())

# Encode file in ASCII characters to send by email    
encoders.encode_base64(part)

# Add header as key/value pair to attachment part
part.add_header(
    "Content-Disposition",
    f"attachment; filename= {filename}",
)

# Add attachment to message and convert message to string
message.attach(part)
text = message.as_string()

# Log in to server using secure context and send email
context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_email, text)